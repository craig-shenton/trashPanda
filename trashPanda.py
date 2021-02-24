# trashPanda function
import os
import yaml
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from fnmatch import fnmatch


def trashPanda(path, config):
    """Takes root directory and YAML config file. Searches root directory for all unstructured excel files, extracts data and returns structured pandas dataframe."""
    my_filenames = [
        os.path.join(root, name)
        for root, dirs, files in os.walk(path)
        for name in files
        if name.endswith((".xlsx"))
    ]
    with open(config, "r") as yamlfile:
        configs = yaml.load(yamlfile, Loader=yaml.FullLoader)
    cols = [
        "file",
        "mod_date",
        "sheet_name",
    ] + configs["Column Names"]
    df = pd.DataFrame(columns=cols)
    for filename in my_filenames:
        wb = load_workbook(filename)
        sheet_list = [
            sheet for sheet in wb.sheetnames if fnmatch(sheet, configs["Sheet"])
        ]
        # last modified date
        mod_date = datetime.fromtimestamp(os.path.getmtime(filename)).strftime(
            "%Y-%m-%d"
        )
        for sheet in sheet_list:
            df = df.append(
                {
                    "file": filename,
                    "mod_date": mod_date,
                    "sheet_name": wb[sheet].title,
                    "A": wb[sheet]["B1"].value,
                    "B": wb[sheet]["D1"].value,
                    "C": wb[sheet]["F1"].value,
                    "D": wb[sheet]["H1"].value,
                    "E": wb[sheet]["B2"].value,
                    "F": wb[sheet]["E2"].value,
                    "Gt": wb[sheet]["G2"].value,
                    "H": wb[sheet]["B3"].value,
                    "I": wb[sheet]["D3"].value,
                    "J": wb[sheet]["F3"].value,
                    "K": wb[sheet]["H3"].value,
                },
                ignore_index=True,
            )
    return df